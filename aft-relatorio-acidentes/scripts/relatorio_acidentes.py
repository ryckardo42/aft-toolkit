# -*- coding: utf-8 -*-
"""
relatorio_acidentes.py - Relatorio de Acidentes do Trabalho (CATs) de um CNPJ.

Skill /aft-relatorio-acidentes do AFT Toolkit. TODO o processamento e local:
os arquivos de CAT contem dados sensiveis (nome, CPF, lesao de trabalhadores)
e NUNCA saem desta maquina. Este script le a fonte, monta o relatorio em
Markdown e em .docx (padrao visual do toolkit) e imprime na tela SOMENTE
numeros agregados e caminhos - nenhum nome de trabalhador aparece no chat.

Dois modos de origem dos dados:

  MODO A - CSV exportado do Portal AFT (arquivo CatsCNPJ_<cnpj>.csv):
      python relatorio_acidentes.py --csv "<arquivo.csv>" --saida "<pasta>"

  MODO B - Base estadual de CATs (planilhas .xlsx, uma por ano, fonte eSocial):
      python relatorio_acidentes.py --cnpj 00.000.000/0000-00 --saida "<pasta>"
      A pasta das planilhas e, por convencao do toolkit, <PASTA_AFT>/CATs -
      ao lado de "OS ATIVAS". Nao precisa configurar nada: basta baixar as
      planilhas do estado (ENIT) e por ali. Quem mantem a base em outro lugar
      usa --base, ou grava o caminho uma vez com --definir-base (campo
      `pasta_cats:` do aft-config.md), que entao prevalece.

  Utilitarios:
      --definir-base "<pasta>"   grava `pasta_cats:` no aft-config.md e sai
      --mostrar-base             mostra a pasta em uso e sai

Saida: Relatorio-Acidentes-<cnpj>.md e .docx na pasta indicada em --saida.
Se os arquivos ja existirem, uma copia .bak-<data-hora> e feita antes.
"""

try:  # ticket automatico de erro (ver _scripts/erro_ticket.py e a skill /aft-erro)
    import sys as _sys
    from pathlib import Path as _Path
    _aqui = _Path(__file__).resolve()
    for _p in (_aqui.parent, *(_a / "_scripts" for _a in _aqui.parents)):
        if (_p / "erro_ticket.py").is_file():
            _sys.path.insert(0, str(_p))
            from erro_ticket import ativar as _ativar_ticket
            _ativar_ticket(__file__)
            break
except Exception:
    pass

import argparse
import csv
import io
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

try:  # console do Windows e cp1252; nunca deixar um acento derrubar o script
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

NAO_INFORMADO = "Não informado"

# eSocial S-2210, campo tpAcid (o CSV do Portal AFT traz so o numero)
TIPO_ACIDENTE = {"1": "Típico", "2": "Doença", "3": "Trajeto"}


# ---------------------------------------------------------------------------
# Localizacao dos modulos do toolkit (pasta_aft, modelo_docx)
# ---------------------------------------------------------------------------
def _skills_raizes():
    """Pastas onde as outras skills do toolkit podem estar: ao lado desta
    (repositorio ou instalacao em ~/.claude/skills) e na instalacao padrao."""
    aqui = Path(__file__).resolve()
    raizes = [aqui.parent.parent.parent,  # <raiz>/aft-relatorio-acidentes/scripts
              Path.home() / ".claude" / "skills"]
    return [r for i, r in enumerate(raizes) if r.is_dir() and r not in raizes[:i]]


def _importar_modulo(subcaminho, nome):
    for raiz in _skills_raizes():
        cand = raiz / subcaminho
        if (cand / (nome + ".py")).is_file():
            sys.path.insert(0, str(cand))
            return __import__(nome)
    return None


def _config_path():
    pa = _importar_modulo("_scripts", "pasta_aft")
    if pa is None:
        return None
    return pa.pasta_aft() / "aft-config.md"


# ---------------------------------------------------------------------------
# Onde ficam as planilhas estaduais
#
# Convencao do toolkit: <PASTA_AFT>/CATs, ao lado de "OS ATIVAS" - assim a base
# acompanha a pasta de trabalho do AFT (que ele pode mover) e ninguem precisa
# configurar caminho nenhum. O campo `pasta_cats:` do aft-config.md continua
# valendo para quem guarda a base em outro lugar, e prevalece quando aponta
# para uma pasta que existe (caminho velho e stale nao derruba a convencao).
# ---------------------------------------------------------------------------
NOMES_PASTA_CATS = ("CATs", "CAT")


def base_no_config():
    """Caminho gravado em `pasta_cats:`, exista ele ou nao (None se ausente)."""
    cfg = _config_path()
    if not (cfg and cfg.is_file()):
        return None
    for linha in cfg.read_text(encoding="utf-8").splitlines():
        m = re.match(r'\s*pasta_cats\s*:\s*"?([^"#]+?)"?\s*$', linha)
        if m and m.group(1).strip():
            return Path(m.group(1).strip()).expanduser()
    return None


def base_padrao():
    """<PASTA_AFT>/CATs, se existir (aceita tambem a variante 'CAT')."""
    pa = _importar_modulo("_scripts", "pasta_aft")
    if pa is None:
        return None
    try:
        raiz = Path(pa.pasta_aft())
    except Exception:
        return None
    for nome in NOMES_PASTA_CATS:
        if (raiz / nome).is_dir():
            return raiz / nome
    return None


def base_configurada():
    """Pasta em uso: o `pasta_cats:` quando aponta para pasta existente; senao
    a convencao <PASTA_AFT>/CATs."""
    cfg = base_no_config()
    if cfg and cfg.is_dir():
        return cfg
    return base_padrao()


# Area do ENIT no SharePoint do MTE com as planilhas de CAT do eSocial por UF.
# So abre com a conta institucional (Microsoft) do auditor ja logada.
LINK_ENIT_CATS = (
    "https://mtegovbr-my.sharepoint.com/shared?id=%2Fpersonal%2Fjoao%5Freis"
    "%5Ftrabalho%5Fgov%5Fbr%2FDocuments%2FDados%2FCATs%20eSocial%20por%20UF"
    "&listurl=%2Fpersonal%2Fjoao%5Freis%5Ftrabalho%5Fgov%5Fbr%2FDocuments"
    "&viewid=68794266%2Df39f%2D4837%2D9e12%2Ddd5cbd44066e&ga=1")


def instrucao_montar_base():
    """Texto unico de 'como montar a base' - o script fala, a skill so repassa."""
    pa = _importar_modulo("_scripts", "pasta_aft")
    try:
        alvo = str(Path(pa.pasta_aft()) / "CATs") if pa else "<pasta AFT>/CATs"
    except Exception:
        alvo = "<pasta AFT>/CATs"
    return (
        f"Crie a pasta  {alvo}  (ao lado de \"OS ATIVAS\") e coloque nela as\n"
        "planilhas .xlsx de CAT do seu estado - uma por ano, todas as que houver.\n"
        "Elas ficam na area do ENIT no SharePoint do MTE, em \"CATs eSocial por UF\":\n"
        f"{LINK_ENIT_CATS}\n"
        "O link so abre com a sua conta institucional (Microsoft) ja logada; entre na\n"
        "pasta da SUA UF e baixe todas as planilhas. Feito isso, nada mais precisa ser\n"
        "configurado.\n"
        "Sem elas nao ha onde procurar: o relatorio de acidentes nao sai, e a preparacao\n"
        "da acao fiscal monta o dossie sem os ultimos acidentes da empresa.\n"
        "Atalho: se o seu Gmail e autorizado do notebooks-aft, o /aft-setup (Passo 2a)\n"
        "baixa e atualiza essas planilhas sozinho (script sincronizar_cats.py).\n"
        "Se preferir manter a base em outro lugar, rode: --definir-base \"<pasta>\"")


def definir_base(pasta):
    pasta = Path(pasta).expanduser()
    if not pasta.is_dir():
        return {"ok": False, "erro": f"pasta não existe: {pasta}"}
    if not list(pasta.glob("*.xlsx")):
        return {"ok": False, "erro": f"nenhum .xlsx encontrado em: {pasta}"}
    cfg = _config_path()
    if cfg is None or not cfg.is_file():
        return {"ok": False, "erro": "aft-config.md não encontrado - rode /aft-setup primeiro"}
    texto = cfg.read_text(encoding="utf-8")
    linha_nova = 'pasta_cats: "%s"' % str(pasta)
    if re.search(r'^\s*pasta_cats\s*:', texto, flags=re.MULTILINE):
        texto = re.sub(r'^\s*pasta_cats\s*:.*$', linha_nova, texto,
                       count=1, flags=re.MULTILINE)
    else:
        bloco = ("\n# Base estadual de Comunicacoes de Acidente (planilhas .xlsx, uma por ano),\n"
                 "# usada pela skill /aft-relatorio-acidentes (Modo B):\n" + linha_nova + "\n")
        texto = texto.rstrip() + "\n" + bloco
    cfg.write_text(texto, encoding="utf-8")
    return {"ok": True, "pasta_cats": str(pasta), "config": str(cfg)}


# ---------------------------------------------------------------------------
# Limpeza de campo (formatacao Excel, mojibake, espacos)
# ---------------------------------------------------------------------------
_GHOST = {  # bytes acentuados descartados em texto MAIUSCULO (base CATWEB antiga)
    "PRODU AO": "PRODUÇÃO", "OPERA AO": "OPERAÇÃO", "ROTULA AO": "ROTULAÇÃO",
    "INSTALA AO": "INSTALAÇÃO", "MANUTEN AO": "MANUTENÇÃO",
    "CONSTRU AO": "CONSTRUÇÃO", "INFORMA AO": "INFORMAÇÃO",
}


def _des_excel(s):
    """Remove a formatacao de exportacao Excel: ="...", "...", =..., aspas."""
    anterior = None
    while s != anterior:
        anterior = s
        s = s.strip()
        for pat in (r'^="(.*)"$', r'^"(.*)"$', r'^=(.*)$'):
            m = re.match(pat, s, flags=re.DOTALL)
            if m:
                s = m.group(1)
                break
    return s


def _des_mojibake(s):
    """Conserta os dois defeitos de codificacao vistos nas bases reais:
    C) 'ø' no lugar de 'Ã' em texto maiusculo (ROTULAÇøO -> ROTULAÇÃO);
    A) dupla codificacao UTF-8 lida como Latin-1 (AÃ\x87OUGUEIRO -> AÇOUGUEIRO)."""
    if "ø" in s and re.search(r"[A-ZÀ-ÖÙ-Þ]ø|ø[A-ZÀ-ÖÙ-Þ]", s):
        s = s.replace("ø", "Ã")
    if "Ã" in s or "Â" in s:
        for cod in ("latin-1", "cp1252"):
            try:
                s2 = s.encode(cod).decode("utf-8")
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
            if s2.count("Ã") < s.count("Ã"):
                s = s2
            break
    return s


def limpar(valor):
    if valor is None:
        return ""
    s = str(valor)
    s = _des_excel(s)
    s = _des_mojibake(s)
    s = re.sub(r"[\r\n]+", " ", s)
    s = re.sub(r"\s{2,}", " ", s).strip()
    if s and s == s.upper():
        for errado, certo in _GHOST.items():
            s = s.replace(errado, certo)
    return s


def so_digitos(s):
    return re.sub(r"\D", "", str(s or ""))


def cnpj_fmt(digitos):
    d = so_digitos(digitos).zfill(14)
    return f"{d[0:2]}.{d[2:5]}.{d[5:8]}/{d[8:12]}-{d[12:14]}"


def _data(valor):
    """datetime a partir de 'dd/mm/aaaa', 'aaaa-mm-dd' ou celula de data."""
    if isinstance(valor, datetime):
        return valor
    s = limpar(valor)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _data_fmt(valor):
    d = _data(valor)
    if d:
        return d.strftime("%d/%m/%Y")
    return limpar(valor) or NAO_INFORMADO


# ---------------------------------------------------------------------------
# MODO A - CSV do Portal AFT
# ---------------------------------------------------------------------------
def ler_csv_portal(caminho):
    caminho = Path(caminho)
    bruto = caminho.read_text(encoding="utf-8-sig", errors="replace")
    linhas = bruto.splitlines()

    empresa, cnpj = "", ""
    idx_cab = None
    for i, linha in enumerate(linhas):
        if linha.startswith("IdAcidente"):
            idx_cab = i
            break
        celulas = [limpar(c) for c in linha.split(";")]
        for j, c in enumerate(celulas):
            if c.startswith("Inscrição") and j + 1 < len(celulas):
                cnpj = so_digitos(celulas[j + 1]) or cnpj
            if c.startswith("Empregador") and j + 1 < len(celulas):
                empresa = celulas[j + 1] or empresa
    if idx_cab is None:
        raise SystemExit("ERRO: cabeçalho 'IdAcidente' não encontrado - este arquivo "
                         "não parece ser o CSV de CATs do Portal AFT.")

    leitor = csv.reader(io.StringIO("\n".join(linhas[idx_cab:])), delimiter=";")
    cab = next(leitor)
    pos = {nome: i for i, nome in enumerate(cab)}

    def campo(row, nome):
        i = pos.get(nome)
        return limpar(row[i]) if i is not None and i < len(row) else ""

    registros = []
    for row in leitor:
        if not any(str(c).strip() for c in row):
            continue
        registros.append({
            "dt": campo(row, "DtAcidente"),
            "trab": campo(row, "NomeTrab"),
            "cargo": campo(row, "NomeCargo"),
            "adm": campo(row, "DtAdmissao"),
            "lesao": campo(row, "DsLesao"),
            "parte": campo(row, "DsParteAtingida"),
            "complesao": campo(row, "DsComplesao"),
            "cid": campo(row, "DsCid"),
            "causa": campo(row, "DsAgenteCausador"),
            "local": campo(row, "DsLocalAcidente"),
            "tratamento": campo(row, "DuracaoTratamento"),
            "municipio": " - ".join(v for v in (campo(row, "NoMunicipioLocalAcidente"),
                                                campo(row, "SgUfLocalAcidente")) if v),
            "obs": campo(row, "ObsCat"),
            "obito": campo(row, "IndCatObito"),
            "dt_obito": campo(row, "DtObito"),
            "tipo_acidente": TIPO_ACIDENTE.get(campo(row, "TpAcidente"),
                                               campo(row, "TpAcidente")),
            "tipo_cat": campo(row, "DsTipoCat"),
            "origem": campo(row, "DsOrigemCat"),
            "ano_base": "",
        })
    fonte = f"CSV do Portal AFT ({caminho.name})"
    return empresa, cnpj, registros, fonte


# ---------------------------------------------------------------------------
# MODO B - base estadual (planilhas .xlsx do eSocial, uma por ano)
# ---------------------------------------------------------------------------
COL_FILTRO = "Número de inscrição do estabelecimento onde o trabalhador exerce atividades"

COLS_B = {  # campo do relatorio -> nome exato da coluna nas planilhas
    "dt": "Data do acidente",
    "trab": "Nome do trabalhador",
    "cargo": "Nome do cargo",
    "adm": "Data de admissão",
    "lesao": "Natureza da lesão",
    "parte": "Parte do corpo atingida",
    "complesao": "Descrição complementar da lesão",
    "cid": "CID",
    "causa": "Agente causador",
    "local": "Especificação do local do acidente",
    "municipio_a": "Município do local do acidente",
    "municipio_b": "UF do local do acidente",
    "tratamento": "Duração estimada do tratamento, em dias",
    "obs": "Observações da CAT",
    "obito": "Indicativo de óbito",
    "dt_obito": "Data do óbito",
    "tipo_acidente": "Tipo de acidente",
    "tipo_cat": "Tipo de CAT",                 # so nas planilhas mais novas
    "razao": "Razão social do empregador",
    "recibo": "Número do recibo",              # idem
    "retif": "Indicativo de retificação",      # idem
    "recibo_retif": "Número do recibo do arquivo retificado",  # idem
}


def ler_base_estadual(pasta, cnpj_alvo):
    try:
        import openpyxl
    except ModuleNotFoundError:
        raise SystemExit("ERRO: falta a biblioteca 'openpyxl' neste Python. "
                         "Instale com: <python> -m pip install openpyxl")

    alvo = so_digitos(cnpj_alvo).zfill(14)
    arquivos = sorted(p for p in Path(pasta).glob("*.xlsx")
                      if not p.name.startswith("~$"))
    if not arquivos:
        raise SystemExit(f"ERRO: nenhum .xlsx encontrado em {pasta}")

    registros, razoes = [], {}
    lidos = []
    for arq in arquivos:
        m = re.search(r"(20\d\d)", arq.stem)
        ano = m.group(1) if m else arq.stem
        wb = openpyxl.load_workbook(arq, read_only=True)
        ws = wb[wb.sheetnames[0]]
        linhas = ws.iter_rows(values_only=True)
        pos = None
        for row in linhas:  # o cabecalho fica apos 2 linhas de titulo
            celulas = [str(c).strip() if c is not None else "" for c in row]
            if COL_FILTRO in celulas:
                pos = {nome: i for i, nome in enumerate(celulas)}
                break
        if pos is None:
            wb.close()
            print(f"AVISO: '{arq.name}' ignorado - não tem a coluna de inscrição "
                  f"do estabelecimento (não parece a planilha estadual de CATs).")
            continue

        i_filtro = pos[COL_FILTRO]
        n_arq = 0
        for row in linhas:
            insc = row[i_filtro] if i_filtro < len(row) else None
            if so_digitos(insc).zfill(14) != alvo:
                continue

            def campo(nome):
                i = pos.get(COLS_B.get(nome, ""))
                if i is None or i >= len(row):
                    return ""
                v = row[i]
                if isinstance(v, datetime):
                    return v.strftime("%d/%m/%Y")
                return limpar(v)

            reg = {chave: campo(chave) for chave in COLS_B}
            reg["municipio"] = " - ".join(v for v in (reg.pop("municipio_a"),
                                                      reg.pop("municipio_b")) if v)
            reg["origem"] = "eSocial"
            reg["ano_base"] = ano
            registros.append(reg)
            if reg.get("razao"):
                razoes[reg["razao"]] = razoes.get(reg["razao"], 0) + 1
            n_arq += 1
        wb.close()
        lidos.append(f"{arq.name}: {n_arq}")

    # Retificacao substitui o registro original: descarta a CAT retificada
    substituidos = {r["recibo_retif"] for r in registros if r.get("recibo_retif")}
    descartados = 0
    if substituidos:
        antes = len(registros)
        registros = [r for r in registros
                     if not (r.get("recibo") and r["recibo"] in substituidos)]
        descartados = antes - len(registros)

    empresa = max(razoes, key=razoes.get) if razoes else ""
    fonte = "base estadual de CATs (eSocial) - " + "; ".join(lidos)
    return empresa, alvo, registros, fonte, descartados


# ---------------------------------------------------------------------------
# Montagem do relatorio
# ---------------------------------------------------------------------------
def _ordenar(registros):
    com, sem = [], []
    for r in registros:
        (com if _data(r.get("dt")) else sem).append(r)
    com.sort(key=lambda r: _data(r["dt"]))
    return com + sem


def _eh_obito(r):
    return (r.get("obito") == "S" or bool(_data(r.get("dt_obito")))
            or r.get("tipo_cat") == "Comunicação de óbito")


def _rotulo(i, r):
    extras = []
    if _eh_obito(r):
        extras.append("ÓBITO")
    if r.get("tipo_cat") and r["tipo_cat"] != "Inicial":
        extras.append(r["tipo_cat"])
    if r.get("retif") == "Retificação":
        extras.append("Retificação")
    sufixo = " (" + " - ".join(extras) + ")" if extras else ""
    return f"Acidente {i}{sufixo}:"


def _ou_ni(v):
    return v if v else NAO_INFORMADO


def _linhas_acidente(r):
    """Pares (rotulo, valor) de um acidente - mesma ordem no MD e no .docx."""
    lesao = _ou_ni(r.get("lesao"))
    if r.get("parte"):
        lesao += f" ({r['parte']})"
    lesao += f" — CID: {_ou_ni(r.get('cid'))}"
    if r.get("complesao"):
        lesao += f' - "{r["complesao"]}"'
    trat = r.get("tratamento")
    trat = f"{trat} dias" if trat and so_digitos(trat) == trat and trat != "0" \
        else _ou_ni("" if trat in ("", "0") else trat)
    pares = [
        ("Dia", _data_fmt(r.get("dt"))),
        ("Trabalhador", _ou_ni(r.get("trab"))),
        ("Cargo", _ou_ni(r.get("cargo"))),
        ("Adm", _data_fmt(r.get("adm"))),
        ("Tipo", _ou_ni(r.get("tipo_acidente"))),
        ("Lesão", lesao),
        ("Causa", _ou_ni(r.get("causa"))),
        ("Local", r.get("local") or
         (f"Município: {r['municipio']}" if r.get("municipio") else NAO_INFORMADO)),
        ("Tratamento", trat),
    ]
    if _data(r.get("dt_obito")):
        pares.append(("Data do óbito", _data_fmt(r["dt_obito"])))
    if r.get("obs"):
        pares.append(("Obs. CAT", r["obs"]))
    return pares


def _estatisticas(registros):
    e = {"total": len(registros), "obitos": sum(1 for r in registros if _eh_obito(r))}
    datas = [_data(r["dt"]) for r in registros if _data(r.get("dt"))]
    e["periodo"] = (f"{min(datas).strftime('%d/%m/%Y')} a "
                    f"{max(datas).strftime('%d/%m/%Y')}") if datas else NAO_INFORMADO
    for chave, campo in (("tipos", "tipo_acidente"), ("cats", "tipo_cat"),
                         ("anos", "ano_base"), ("causas", "causa"),
                         ("partes", "parte")):
        cont = {}
        for r in registros:
            v = r.get(campo)
            if v:
                cont[v] = cont.get(v, 0) + 1
        e[chave] = cont
    if set(e["cats"]) <= {"Inicial"}:  # so vale a pena mostrar se ha reabertura etc.
        e["cats"] = {}
    return e


def _top(contagem, n=5):
    """Os n itens mais frequentes, como 'Nome (12) · Nome (8) · ...'.
    Rotulos do eSocial podem ser paragrafos inteiros: corta no resumo
    (a descricao completa continua na listagem por acidente)."""
    itens = sorted(contagem.items(), key=lambda kv: (-kv[1], kv[0]))[:n]
    return " · ".join(f"{k[:57] + '...' if len(k) > 60 else k} ({v})"
                      for k, v in itens)


def gerar_md(caminho, empresa, cnpj14, registros, fonte, est):
    hoje = datetime.now().strftime("%d/%m/%Y")
    L = ["# Relatório de Acidentes do Trabalho (CATs)", ""]
    L += [f"**Empresa:** {_ou_ni(empresa)}",
          f"**CNPJ:** {cnpj_fmt(cnpj14)}",
          f"**Fonte:** {fonte}",
          f"**Gerado em:** {hoje}", ""]
    L += ["## Resumo", "",
          f"- Total de CATs: {est['total']}",
          f"- CATs com óbito: {est['obitos']}",
          f"- Período dos acidentes: {est['periodo']}"]
    if est["tipos"]:
        L.append("- Por tipo: " + " · ".join(f"{k}: {v}" for k, v in
                                             sorted(est["tipos"].items())))
    if est["cats"]:
        L.append("- Por tipo de CAT: " + " · ".join(f"{k}: {v}" for k, v in
                                                    sorted(est["cats"].items())))
    if est["anos"]:
        L.append("- Por planilha (ano-base): " + " · ".join(
            f"{k}: {v}" for k, v in sorted(est["anos"].items())))
    if est["causas"]:
        L.append("- Principais agentes causadores: " + _top(est["causas"]))
    if est["partes"]:
        L.append("- Partes do corpo mais atingidas: " + _top(est["partes"]))
    L += ["", "## Acidentes", ""]
    for i, r in enumerate(registros, 1):
        L.append(f"* **{_rotulo(i, r)}**")
        for rotulo, valor in _linhas_acidente(r):
            L.append(f"    * **{rotulo}:** {valor}")
        L.append("")
    L += ["---", f"Total de acidentes registrados: {est['total']}"]
    if est["obitos"]:
        L.append(f"Óbitos: {est['obitos']}")
    Path(caminho).write_text("\n".join(L) + "\n", encoding="utf-8")


def gerar_docx(caminho, empresa, cnpj14, registros, fonte, est):
    m = _importar_modulo("aft-modelo-docx/scripts", "modelo_docx")
    if m is None:
        return False
    hoje = datetime.now().strftime("%d/%m/%Y")
    doc = m.novo_documento()
    m.capa(doc, "RELATÓRIO DE ACIDENTES DO TRABALHO",
           subtitulo="Comunicações de Acidente de Trabalho (CAT) registradas",
           unidade=f"{_ou_ni(empresa)} — CNPJ {cnpj_fmt(cnpj14)}",
           data=f"Gerado em {hoje}")

    m.titulo_secao(doc, "1. Resumo")
    linhas = [("Total de CATs", str(est["total"])),
              ("CATs com óbito", str(est["obitos"])),
              ("Período dos acidentes", est["periodo"])]
    if est["tipos"]:
        linhas.append(("Por tipo de acidente", " · ".join(
            f"{k}: {v}" for k, v in sorted(est["tipos"].items()))))
    if est["cats"]:
        linhas.append(("Por tipo de CAT", " · ".join(
            f"{k}: {v}" for k, v in sorted(est["cats"].items()))))
    if est["anos"]:
        linhas.append(("Por planilha (ano-base)", " · ".join(
            f"{k}: {v}" for k, v in sorted(est["anos"].items()))))
    if est["causas"]:
        linhas.append(("Principais agentes causadores", _top(est["causas"])))
    if est["partes"]:
        linhas.append(("Partes do corpo mais atingidas", _top(est["partes"])))
    linhas.append(("Fonte", fonte))
    m.tabela_rotulo_valor(doc, linhas)

    m.titulo_secao(doc, "2. Relação de acidentes")
    for i, r in enumerate(registros, 1):
        m.subtitulo(doc, f"{_rotulo(i, r)[:-1]} — {_data_fmt(r.get('dt'))}")
        for rotulo, valor in _linhas_acidente(r):
            if rotulo == "Dia":
                continue  # ja esta no subtitulo
            m.marcador(doc, f"{rotulo}: {valor}")

    m.paragrafo(doc)
    m.paragrafo(doc, "Documento de trabalho gerado pelo AFT Toolkit "
                     "(/aft-relatorio-acidentes) a partir dos registros de CAT. "
                     "Os dados reproduzem o que foi declarado nas comunicações "
                     "e estão sujeitos a conferência pelo Auditor-Fiscal.",
                italico=True)
    doc.save(str(caminho))
    return True


def _backup(caminho):
    if Path(caminho).exists():
        carimbo = datetime.now().strftime("%Y%m%d-%H%M%S")
        destino = str(caminho) + ".bak-" + carimbo
        shutil.copy2(caminho, destino)
        return destino
    return None


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(description="Relatorio de Acidentes (CATs) de um CNPJ")
    ap.add_argument("--csv", help="Modo A: CSV de CATs exportado do Portal AFT")
    ap.add_argument("--cnpj", help="CNPJ do estabelecimento (obrigatorio no Modo B)")
    ap.add_argument("--base", help="Modo B: pasta das planilhas estaduais (senao usa o aft-config.md)")
    ap.add_argument("--saida", help="pasta onde gravar o relatorio (MD + DOCX)")
    ap.add_argument("--empresa", help="razao social a usar no relatorio (opcional)")
    ap.add_argument("--definir-base", help="grava a pasta das planilhas no aft-config.md e sai")
    ap.add_argument("--mostrar-base", action="store_true", help="mostra a pasta configurada e sai")
    args = ap.parse_args()

    if args.definir_base:
        r = definir_base(args.definir_base)
        print("OK: pasta_cats gravada em %s -> %s" % (r.get("config"), r.get("pasta_cats"))
              if r["ok"] else "ERRO: " + r["erro"])
        sys.exit(0 if r["ok"] else 1)

    if args.mostrar_base:
        base, cfg = base_configurada(), base_no_config()
        if base:
            origem = "aft-config.md" if cfg and cfg.is_dir() else "convenção <PASTA_AFT>/CATs"
            print(f"pasta_cats: {base}   [{origem}]")
            print(f"planilhas: {len(list(base.glob('*.xlsx')))} .xlsx")
            if cfg and not cfg.is_dir():
                print(f"AVISO: o aft-config.md aponta para uma pasta que não existe "
                      f"({cfg}) — ignorada em favor da convenção. Apague a linha "
                      "pasta_cats: do aft-config.md para não confundir.")
        else:
            print("PASTA_CATS_NAO_DEFINIDA")
            print(instrucao_montar_base())
        sys.exit(0)

    if not args.saida:
        ap.error("--saida é obrigatório")

    descartados = 0
    if args.csv:  # -------------------------------------------------- Modo A
        empresa, cnpj14, registros, fonte = ler_csv_portal(args.csv)
        if args.cnpj and so_digitos(args.cnpj).zfill(14) != so_digitos(cnpj14).zfill(14):
            raise SystemExit(f"ERRO: o CSV é do CNPJ {cnpj_fmt(cnpj14)}, não do "
                             f"CNPJ informado ({cnpj_fmt(args.cnpj)}). Confira o arquivo.")
        modo = "A (CSV do Portal AFT)"
    else:  # ---------------------------------------------------------- Modo B
        if not args.cnpj:
            ap.error("informe --csv (Modo A) ou --cnpj (Modo B)")
        base = Path(args.base).expanduser() if args.base else base_configurada()
        if not base:
            print("PASTA_CATS_NAO_DEFINIDA: não encontrei a base estadual de CATs.")
            print(instrucao_montar_base())
            sys.exit(2)
        if not base.is_dir():
            raise SystemExit(f"ERRO: a pasta configurada não existe: {base}")
        empresa, cnpj14, registros, fonte, descartados = ler_base_estadual(base, args.cnpj)
        modo = f"B (base estadual: {base})"

    if args.empresa:
        empresa = args.empresa

    if not registros:
        print(f"NENHUMA_CAT: não há CAT para o CNPJ {cnpj_fmt(cnpj14)} na fonte "
              f"consultada ({modo}). Nenhum arquivo foi gerado.")
        sys.exit(3)

    registros = _ordenar(registros)
    est = _estatisticas(registros)

    saida = Path(args.saida).expanduser()
    saida.mkdir(parents=True, exist_ok=True)
    nome = "Relatorio-Acidentes-" + so_digitos(cnpj14).zfill(14)
    md_path = saida / (nome + ".md")
    docx_path = saida / (nome + ".docx")

    backups = [b for b in (_backup(md_path), _backup(docx_path)) if b]
    gerar_md(md_path, empresa, cnpj14, registros, fonte, est)
    docx_ok = gerar_docx(docx_path, empresa, cnpj14, registros, fonte, est)

    # Resumo para o chat: SOMENTE agregados e caminhos - nunca nome de trabalhador.
    print("RELATORIO_GERADO")
    print(f"  Modo: {modo}")
    print(f"  CNPJ: {cnpj_fmt(cnpj14)}")
    print(f"  Empresa: {_ou_ni(empresa)}")
    print(f"  Total de CATs: {est['total']} | Com óbito: {est['obitos']} | "
          f"Período: {est['periodo']}")
    if est["tipos"]:
        print("  Por tipo: " + " · ".join(f"{k}: {v}" for k, v in sorted(est["tipos"].items())))
    if est["anos"]:
        print("  Por planilha: " + " · ".join(f"{k}: {v}" for k, v in sorted(est["anos"].items())))
    if est["causas"]:
        print("  Principais agentes causadores: " + _top(est["causas"], 3))
    if est["partes"]:
        print("  Partes do corpo mais atingidas: " + _top(est["partes"], 3))
    if descartados:
        print(f"  CATs retificadas descartadas (substituídas pela retificação): {descartados}")
    print(f"  MD:   {md_path}")
    if docx_ok:
        print(f"  DOCX: {docx_path}")
    else:
        print("  DOCX_FALHOU: a skill aft-modelo-docx não foi encontrada - só o MD foi gerado.")
    for b in backups:
        print(f"  Backup do arquivo anterior: {b}")


if __name__ == "__main__":
    main()
