# -*- coding: utf-8 -*-
"""
cat_trabalhador.py - Dossie de CATs de UM trabalhador (por CPF ou nome).

Skill /aft-cat-trabalhador do AFT Toolkit. TODO o processamento e local: as
planilhas de CAT contem nome, CPF e dados de saude de trabalhadores, e nada
disso sai da maquina. O script varre a base estadual de CATs (as mesmas
planilhas .xlsx da /aft-relatorio-acidentes), junta todas as CATs do
trabalhador e gera um PDF no leiaute do formulario CAT do eSocial - uma ficha
completa por CAT, em ordem cronologica. Na tela sai apenas o nome, o
CPF MASCARADO e numeros agregados.

Uso:
    python cat_trabalhador.py --cpf 00000000000 --saida "<pasta>"
    python cat_trabalhador.py --nome "FULANO DA SILVA" --saida "<pasta>"
    python cat_trabalhador.py --nome "FULANO" --indice 2 --saida "<pasta>"
                              (quando a busca por nome achou mais de uma pessoa)
    python cat_trabalhador.py --mostrar-base

A pasta das planilhas e a mesma da /aft-relatorio-acidentes: por convencao
<PASTA_AFT>/CATs, ou o campo `pasta_cats:` do aft-config.md; --base sobrepoe.

Saida: Dossie-CAT-<cpf>.pdf na pasta indicada em --saida (backup .bak-<data>
se ja existir). Requer openpyxl (leitura) e reportlab (PDF).
"""

try:  # ticket automatico de erro (ver _scripts/erro_ticket.py e a skill /aft-erro)
    import sys as _sys
    from pathlib import Path as _Path
    _aqui = _Path(__file__).resolve()
    for _p in (_aqui.parent, *(_a / "_scripts" for _a in _aqui.parents)):
        if (_p / "erro_ticket.py").is_file():
            _sys.path.insert(0, str(_p))
            from erro_ticket import ativar as _ativar_ticket
            _ativar_ticket(__file__)
            break
except Exception:
    pass

import argparse
import datetime as _dt
import json
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:  # console do Windows e cp1252; nunca deixar um acento derrubar o script
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass


# ---------------------------------------------------------------------------
# Reuso da /aft-relatorio-acidentes: localizacao da base e limpeza de campo
# ---------------------------------------------------------------------------
def _skills_raizes():
    aqui = Path(__file__).resolve()
    raizes = [aqui.parent.parent.parent,  # <raiz>/aft-cat-trabalhador/scripts
              Path.home() / ".claude" / "skills"]
    return [r for i, r in enumerate(raizes) if r.is_dir() and r not in raizes[:i]]


def _importar_modulo(subcaminho, nome):
    for raiz in _skills_raizes():
        cand = raiz / subcaminho
        if (cand / (nome + ".py")).is_file():
            sys.path.insert(0, str(cand))
            return __import__(nome)
    return None


ra = _importar_modulo("aft-relatorio-acidentes/scripts", "relatorio_acidentes")
if ra is None:
    raise SystemExit("ERRO: a skill aft-relatorio-acidentes não foi encontrada - "
                     "este dossiê usa a mesma base de CATs dela. Rode /aft-atualizar.")

limpar = ra.limpar
so_digitos = ra.so_digitos
cnpj_fmt = ra.cnpj_fmt
_data = ra._data
NAO_INFORMADO = "Não informado"
TRACO = "-"


def cpf_fmt(digitos):
    d = so_digitos(digitos).zfill(11)
    return f"{d[0:3]}.{d[3:6]}.{d[6:9]}-{d[9:11]}"


def cpf_mascarado(digitos):
    """Para a tela: só os 3 primeiros e os 2 últimos dígitos (regra do chat)."""
    d = so_digitos(digitos).zfill(11)
    return f"{d[0:3]}.***.***-{d[9:11]}"


def _sem_acento(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c))


def _nome_norm(s):
    return re.sub(r"\s+", " ", _sem_acento(s).upper()).strip()


# ---------------------------------------------------------------------------
# Leitura da base: todas as colunas, filtrando por CPF ou nome
# ---------------------------------------------------------------------------
COL_CPF = "CPF"
COL_NOME = "Nome do trabalhador"


def _valor_celula(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, _dt.time):
        return v.strftime("%H:%M")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return limpar(v)


def ler_base(pasta, cpf_alvo=None, nome_alvo=None):
    """Registros do trabalhador, com TODAS as colunas da planilha (dict
    nome-da-coluna -> valor limpo) + 'Ano da planilha'."""
    try:
        import openpyxl
    except ModuleNotFoundError:
        raise SystemExit("ERRO: falta a biblioteca 'openpyxl' neste Python. "
                         "Instale com: <python> -m pip install openpyxl")

    alvo_cpf = so_digitos(cpf_alvo).zfill(11) if cpf_alvo else None
    alvo_nome = _nome_norm(nome_alvo) if nome_alvo else None

    arquivos = sorted(p for p in Path(pasta).glob("*.xlsx")
                      if not p.name.startswith("~$"))
    if not arquivos:
        raise SystemExit(f"ERRO: nenhum .xlsx encontrado em {pasta}")

    registros, lidos = [], []
    for arq in arquivos:
        m = re.search(r"(20\d\d)", arq.stem)
        ano = m.group(1) if m else arq.stem
        wb = openpyxl.load_workbook(arq, read_only=True)
        ws = wb[wb.sheetnames[0]]
        linhas = ws.iter_rows(values_only=True)
        pos = None
        for row in linhas:  # o cabecalho fica apos 2 linhas de titulo
            celulas = [str(c).strip() if c is not None else "" for c in row]
            if COL_CPF in celulas and COL_NOME in celulas:
                pos = {nome: i for i, nome in enumerate(celulas) if nome}
                break
        if pos is None:
            wb.close()
            print(f"AVISO: '{arq.name}' ignorado - não tem as colunas de CPF e "
                  f"nome do trabalhador (não parece a planilha estadual de CATs).")
            continue

        i_cpf, i_nome = pos[COL_CPF], pos[COL_NOME]
        n_arq = 0
        for row in linhas:
            cpf = so_digitos(row[i_cpf] if i_cpf < len(row) else "").zfill(11)
            nome = limpar(row[i_nome] if i_nome < len(row) else "")
            if alvo_cpf:
                if cpf != alvo_cpf:
                    continue
            else:
                n = _nome_norm(nome)
                if not (n == alvo_nome or alvo_nome in n):
                    continue
            reg = {c: _valor_celula(row[i]) if i < len(row) else ""
                   for c, i in pos.items()}
            reg["Ano da planilha"] = ano
            registros.append(reg)
            n_arq += 1
        wb.close()
        lidos.append(f"{arq.name}: {n_arq}")

    # Retificacao substitui o registro original: descarta a CAT retificada
    substituidos = {r.get("Número do recibo do arquivo retificado")
                    for r in registros
                    if r.get("Número do recibo do arquivo retificado")}
    descartados = 0
    if substituidos:
        antes = len(registros)
        registros = [r for r in registros
                     if not (r.get("Número do recibo")
                             and r["Número do recibo"] in substituidos)]
        descartados = antes - len(registros)

    fonte = "base estadual de CATs (eSocial) - " + "; ".join(lidos)
    return registros, fonte, descartados


def _ordenar(registros):
    com, sem = [], []
    for r in registros:
        (com if _data(r.get("Data do acidente")) else sem).append(r)
    com.sort(key=lambda r: _data(r["Data do acidente"]))
    return com + sem


def agrupar_por_cpf(registros):
    """{cpf11: [registros]} preservando a ordem de descoberta."""
    grupos = {}
    for r in registros:
        cpf = so_digitos(r.get(COL_CPF)).zfill(11)
        grupos.setdefault(cpf, []).append(r)
    return grupos


def _eh_obito(r):
    return (r.get("Indicativo de óbito") == "S"
            or bool(_data(r.get("Data do óbito")))
            or r.get("Tipo de CAT") == "Comunicação de óbito")


# ---------------------------------------------------------------------------
# PDF no leiaute do formulario CAT do eSocial (reportlab)
# ---------------------------------------------------------------------------
NOTA_RODAPE = "Documento gerado a partir dos dados enviados pela empresa ao eSocial."


def _v(r, coluna):
    return r.get(coluna) or TRACO


def _endereco_acidente(r):
    partes = [r.get("Tipo de logradouro do local do acidente"),
              r.get("Descrição do logradouro do local do acidente"),
              r.get("Número do logradouro do local do acidente")]
    end = " ".join(p for p in partes if p)
    extras = [r.get("Complemento do logradouro do local do acidente"),
              r.get("Nome do bairro/distrito do local do acidente"),
              r.get("CEP do local do acidente") or
              r.get("Código de Endereçamento Postal do local do acidente")]
    tudo = ", ".join(x for x in ([end] + extras) if x)
    return tudo or TRACO


def _cbo(r):
    cod, desc = r.get("CBO - código"), r.get("CBO")
    if cod and desc:
        return f"{cod} - {desc}"
    return cod or desc or TRACO


def _cid(r):
    cod, desc = r.get("CID - código"), r.get("CID")
    if cod and desc:
        return f"{cod} - {desc}"
    return cod or desc or TRACO


def _cnae(r, prefixo):
    cod = r.get(prefixo + " - código")
    desc = r.get(prefixo)
    if cod and desc:
        return f"{cod} - {desc}"
    return cod or desc or TRACO


def _medico(r):
    nome = r.get("Nome do médico/dentista que emitiu o atestado")
    orgao = r.get("Órgão de classe")
    num = r.get("Número de inscrição no órgão de classe")
    uf = r.get("UF do órgão de classe")
    reg = " ".join(x for x in (orgao, num) if x)
    return " - ".join(x for x in (nome, reg, uf) if x) or TRACO


def _sim_nao(v):
    v = (v or "").strip().upper()
    if v.startswith("S"):
        return "SIM"
    if v.startswith("N"):
        return "NÃO"
    return v or TRACO


def _secoes_cat(r):
    """(titulo_secao, [(rotulo, valor)]) de uma CAT, na ordem do formulario
    do eSocial. Campos que a base estadual nao traz ficam de fora."""
    ident = [
        ("Emitente", _v(r, "Emitente da CAT")),
        ("Tipo de CAT", _v(r, "Tipo de CAT")),
        ("Iniciativa da CAT", _v(r, "Iniciativa da CAT")),
        ("Fonte do cadastramento", "eSocial (base estadual de CATs)"),
        ("Número do recibo", _v(r, "Número do recibo")),
        ("Data de emissão da CAT", _v(r, "Data de emissão da CAT")),
        ("Planilha de origem (ano)", _v(r, "Ano da planilha")),
    ]
    if r.get("Recibo da última CAT, em caso de reabertura ou comunicação de óbito"):
        ident.append(("Recibo da CAT de origem",
                      r["Recibo da última CAT, em caso de reabertura ou comunicação de óbito"]))
    if r.get("Indicativo de retificação") and \
            r["Indicativo de retificação"] not in ("Original",):
        ident.append(("Retificação", r["Indicativo de retificação"]))
        if r.get("Número do recibo do arquivo retificado"):
            ident.append(("Recibo do arquivo retificado",
                          r["Número do recibo do arquivo retificado"]))

    empregador = [
        ("Razão social / Nome", _v(r, "Razão social do empregador")),
        ("Tipo de inscrição", _v(r, "Tipo de inscrição do empregador")),
        ("Número de inscrição",
         cnpj_fmt(r["Número de inscrição do empregador"])
         if so_digitos(r.get("Número de inscrição do empregador"))
         else TRACO),
        ("Estabelecimento do trabalhador",
         cnpj_fmt(r["Número de inscrição do estabelecimento onde o trabalhador exerce atividades"])
         if so_digitos(r.get("Número de inscrição do estabelecimento onde o trabalhador exerce atividades"))
         else TRACO),
        ("CNAE", _cnae(r, "CNAE do empregador")),
        ("Município / UF", " - ".join(x for x in (r.get("Município do empregador"),
                                                  r.get("UF do empregador")) if x) or TRACO),
        ("ME / EPP", _sim_nao(r.get("Indicativo de ME ou EPP"))),
    ]

    acidentado = [
        ("Nome", _v(r, COL_NOME)),
        ("CPF", cpf_fmt(r[COL_CPF]) if so_digitos(r.get(COL_CPF)) else TRACO),
        ("NIS", _v(r, "NIS")),
        ("Data de nascimento", _v(r, "Data de nascimento")),
        ("Idade na data do acidente", _v(r, "Idade na data do acidente")),
        ("Sexo", _v(r, "Sexo")),
        ("Raça e cor", _v(r, "Raça e cor")),
        ("Grau de instrução", _v(r, "Grau de instrução")),
        ("CBO", _cbo(r)),
        ("Nome do cargo", _v(r, "Nome do cargo")),
        ("Data de admissão", _v(r, "Data de admissão")),
        ("Matrícula na empresa", _v(r, "Matrícula atribuída ao trabalhador pela empresa")),
        ("Categoria do trabalhador", _v(r, "Categoria do trabalhador")),
        ("Salário base (parte fixa)", _v(r, "Salário base do trabalhador, correspondente à parte fixa da remuneração")),
    ]

    acidente = [
        ("Data do acidente", _v(r, "Data do acidente")),
        ("Hora do acidente", _v(r, "Hora do acidente")),
        ("Horas trabalhadas antes do acidente", _v(r, "Horas trabalhadas antes da ocorrência do acidente")),
        ("Tipo", _v(r, "Tipo de acidente")),
        ("Houve afastamento?", _sim_nao(r.get("Indicativo de afastamento do trabalho durante o tratamento"))),
        ("Tipo de local do acidente", _v(r, "Tipo de local do acidente")),
        ("Especificação do local", _v(r, "Especificação do local do acidente")),
        ("Endereço do local", _endereco_acidente(r)),
        ("Município / UF / País", " - ".join(x for x in (
            r.get("Município do local do acidente"),
            r.get("UF do local do acidente"),
            r.get("País do local do acidente")) if x) or TRACO),
        ("Parte do corpo atingida", _v(r, "Parte do corpo atingida")),
        ("Lateralidade", _v(r, "Lateralidade da(s) parte(s) atingida(s)")),
        ("Agente causador", _v(r, "Agente causador")),
        ("Situação geradora", _v(r, "Situação geradora")),
        ("Comunicação à autoridade policial", _sim_nao(r.get("Indicativo de comunicação à autoridade policial"))),
        ("Houve morte?", _sim_nao(r.get("Indicativo de óbito"))),
        ("Data do óbito", _v(r, "Data do óbito")),
    ]
    insc_ocor = r.get("Número de inscrição do estabelecimento onde ocorreu o acidente ou a doença")
    if insc_ocor or r.get("Razão social do estabelecimento onde ocorreu o acidente ou a doença"):
        acidente += [
            ("Estabelecimento onde ocorreu",
             " - ".join(x for x in (
                 cnpj_fmt(insc_ocor) if so_digitos(insc_ocor) else "",
                 r.get("Razão social do estabelecimento onde ocorreu o acidente ou a doença"))
                 if x) or TRACO),
            ("CNAE do estabelecimento onde ocorreu",
             _cnae(r, "CNAE do estabelecimento onde ocorreu o acidente ou a doença")),
        ]
    if r.get("Observações da CAT"):
        acidente.append(("Observações da CAT", r["Observações da CAT"]))

    atestado = [
        ("Data do atendimento", _v(r, "Data do atendimento")),
        ("Hora do atendimento", _v(r, "Hora do atendimento")),
        ("Houve internação?", _sim_nao(r.get("Indicativo de internação"))),
        ("Duração estimada do tratamento (dias)", _v(r, "Duração estimada do tratamento, em dias")),
        ("Afastamento durante o tratamento?", _sim_nao(r.get("Indicativo de afastamento do trabalho durante o tratamento"))),
        ("Natureza da lesão", _v(r, "Natureza da lesão")),
        ("Descrição complementar da lesão", _v(r, "Descrição complementar da lesão")),
        ("Diagnóstico provável", _v(r, "Diagnóstico provável")),
        ("CID", _cid(r)),
        ("Médico/dentista (nome, registro, UF)", _medico(r)),
        ("Observações do atestado",
         r.get("Observações do atestado") or r.get("Obervações do atestado") or TRACO),
    ]

    return [
        ("I - IDENTIFICAÇÃO DA CAT", ident),
        ("II - EMPREGADOR", empregador),
        ("ACIDENTADO", acidentado),
        ("ACIDENTE OU DOENÇA", acidente),
        ("III - INFORMAÇÕES DO ATESTADO MÉDICO", atestado),
    ]


def gerar_pdf(caminho, cpf11, registros):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (BaseDocTemplate, Frame, PageBreak,
                                        PageTemplate, Paragraph, Spacer, Table,
                                        TableStyle)
    except ModuleNotFoundError:
        raise SystemExit("ERRO: falta a biblioteca 'reportlab' neste Python. "
                         "Instale com: <python> -m pip install reportlab")

    r0 = registros[-1]  # dados cadastrais mais recentes
    nome_trab = r0.get(COL_NOME) or NAO_INFORMADO

    est_rotulo = ParagraphStyle("rot", fontName="Helvetica-Bold", fontSize=8,
                                leading=10)
    est_valor = ParagraphStyle("val", fontName="Helvetica", fontSize=8,
                               leading=10)
    est_secao = ParagraphStyle("sec", fontName="Helvetica-Bold", fontSize=9,
                               leading=12)
    est_titulo = ParagraphStyle("tit", fontName="Helvetica-Bold", fontSize=14,
                                leading=17)
    est_nota = ParagraphStyle("nota", fontName="Helvetica-Oblique", fontSize=7,
                              leading=9, textColor=colors.HexColor("#444444"))

    CINZA = colors.HexColor("#d9d9d9")
    CINZA_CLARO = colors.HexColor("#f0f0f0")
    BORDA = colors.HexColor("#666666")
    LARG = 180 * mm

    def tabela_secao(titulo, pares):
        linhas = [[Paragraph(titulo, est_secao), ""]]
        for rot, val in pares:
            linhas.append([Paragraph(rot, est_rotulo),
                           Paragraph(str(val), est_valor)])
        t = Table(linhas, colWidths=[62 * mm, LARG - 62 * mm])
        t.setStyle(TableStyle([
            ("SPAN", (0, 0), (1, 0)),
            ("BACKGROUND", (0, 0), (1, 0), CINZA),
            ("BACKGROUND", (0, 1), (0, -1), CINZA_CLARO),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDA),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        return t

    def rodape(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#666666"))
        canvas.drawCentredString(A4[0] / 2, 10 * mm, f"Página {doc.page}")
        canvas.restoreState()

    doc = BaseDocTemplate(str(caminho), pagesize=A4,
                          leftMargin=15 * mm, rightMargin=15 * mm,
                          topMargin=14 * mm, bottomMargin=16 * mm,
                          title=f"Dossiê de CATs - {nome_trab}")
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height)
    doc.addPageTemplates([PageTemplate(id="pagina", frames=[frame],
                                       onPage=rodape)])

    # Uma ficha completa por CAT, cada uma comecando em pagina propria
    elems = []
    for i, r in enumerate(registros, 1):
        if i > 1:
            elems.append(PageBreak())
        marca = " - ÓBITO" if _eh_obito(r) else ""
        elems.append(Paragraph(
            f"CAT {i} de {len(registros)} - "
            f"{r.get('Data do acidente') or 'sem data'}{marca}", est_titulo))
        elems.append(Spacer(1, 3 * mm))
        for titulo, pares in _secoes_cat(r):
            elems.append(tabela_secao(titulo, pares))
            elems.append(Spacer(1, 3 * mm))
        elems.append(Paragraph(NOTA_RODAPE, est_nota))

    doc.build(elems)


# ---------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser(
        description="Dossie de CATs de um trabalhador (por CPF ou nome)")
    ap.add_argument("--cpf", help="CPF do trabalhador (com ou sem pontuação)")
    ap.add_argument("--nome", help="nome (ou parte do nome) do trabalhador")
    ap.add_argument("--indice", type=int, metavar="N",
                    help="com --nome ambíguo: gera o dossiê do trabalhador N "
                         "da lista MULTIPLOS_TRABALHADORES")
    ap.add_argument("--base", help="pasta das planilhas estaduais (senão usa a "
                                   "mesma configuração da /aft-relatorio-acidentes)")
    ap.add_argument("--saida", help="pasta onde gravar o PDF")
    ap.add_argument("--mostrar-base", action="store_true",
                    help="mostra a pasta das planilhas em uso e sai")
    args = ap.parse_args()

    if args.mostrar_base:
        base = ra.base_configurada()
        if base:
            print(f"pasta_cats: {base}")
            print("planilhas: %d .xlsx" % len(
                [x for x in base.glob("*.xlsx") if not x.name.startswith("~$")]))
        else:
            print("PASTA_CATS_NAO_DEFINIDA")
            print(ra.instrucao_montar_base())
        sys.exit(0)

    if not args.cpf and not args.nome:
        ap.error("informe --cpf ou --nome")
    if args.cpf and args.nome:
        ap.error("informe --cpf OU --nome, não os dois")
    if args.cpf and len(so_digitos(args.cpf)) != 11:
        ap.error(f"CPF deve ter 11 dígitos (recebi {len(so_digitos(args.cpf))})")
    if not args.saida:
        ap.error("--saida é obrigatório")

    base = Path(args.base).expanduser() if args.base else ra.base_configurada()
    if not base:
        print("PASTA_CATS_NAO_DEFINIDA: não encontrei a base estadual de CATs.")
        print(ra.instrucao_montar_base())
        sys.exit(2)
    if not base.is_dir():
        raise SystemExit(f"ERRO: a pasta da base não existe: {base}")

    registros, _fonte, descartados = ler_base(base, cpf_alvo=args.cpf,
                                              nome_alvo=args.nome)
    if not registros:
        alvo = (f"CPF {cpf_mascarado(args.cpf)}" if args.cpf
                else f'nome "{args.nome}"')
        print(f"NENHUM_TRABALHADOR: nenhuma CAT para o {alvo} na base "
              f"consultada ({base}). Nenhum arquivo foi gerado.")
        sys.exit(3)

    grupos = agrupar_por_cpf(registros)
    if len(grupos) > 1:
        # Busca por nome achou mais de uma pessoa: listar e deixar o AFT
        # escolher (--indice N). CPF nunca sai inteiro na tela.
        chaves = sorted(grupos, key=lambda c: grupos[c][0].get(COL_NOME) or "")
        if args.indice:
            if not 1 <= args.indice <= len(chaves):
                raise SystemExit(f"ERRO: --indice deve estar entre 1 e {len(chaves)}")
            cpf11 = chaves[args.indice - 1]
            registros = grupos[cpf11]
        else:
            lista = []
            for i, cpf in enumerate(chaves, 1):
                rs = grupos[cpf]
                empregadores = sorted({r.get("Razão social do empregador") or "?"
                                       for r in rs})
                lista.append({
                    "indice": i,
                    "nome": rs[0].get(COL_NOME) or NAO_INFORMADO,
                    "cpf_mascarado": cpf_mascarado(cpf),
                    "nascimento": rs[0].get("Data de nascimento") or TRACO,
                    "cats": len(rs),
                    "empregadores": empregadores,
                })
            print("MULTIPLOS_TRABALHADORES " +
                  json.dumps(lista, ensure_ascii=False))
            print("Repita o comando acrescentando --indice N para escolher.")
            sys.exit(4)
    else:
        cpf11 = next(iter(grupos))
        registros = grupos[cpf11]

    registros = _ordenar(registros)

    saida = Path(args.saida).expanduser()
    saida.mkdir(parents=True, exist_ok=True)
    pdf_path = saida / f"Dossie-CAT-{cpf11}.pdf"
    backup = ra._backup(pdf_path)
    gerar_pdf(pdf_path, cpf11, registros)

    # Resumo para o chat: nome e agregados; CPF sempre mascarado, nada da CAT.
    obitos = sum(1 for r in registros if _eh_obito(r))
    anos = {}
    for r in registros:
        d = _data(r.get("Data do acidente"))
        if d:
            anos[str(d.year)] = anos.get(str(d.year), 0) + 1
    empregadores = sorted({r.get("Razão social do empregador") or "?"
                           for r in registros})
    print("DOSSIE_GERADO")
    print(f"  Trabalhador: {registros[-1].get(COL_NOME) or NAO_INFORMADO}")
    print(f"  CPF: {cpf_mascarado(cpf11)}")
    print(f"  CATs no dossiê: {len(registros)}" +
          (f" | Com óbito: {obitos}" if obitos else ""))
    if anos:
        print("  Por ano: " + " · ".join(f"{k}: {v}" for k, v in sorted(anos.items())))
    print(f"  Empregador(es): {'; '.join(empregadores)}")
    if descartados:
        print(f"  CATs retificadas descartadas (substituídas pela retificação): {descartados}")
    print(f"  PDF: {pdf_path}")
    if backup:
        print(f"  Backup do arquivo anterior: {backup}")


if __name__ == "__main__":
    main()
